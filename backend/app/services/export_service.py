import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, Series

def generate_dynamic_excel_report(data_list):
    """
    Generate a dynamic Excel report with Native Bar Charts based on live frontend data.
    """
    df = pd.DataFrame(data_list)
    
    # Split the dataset into Raw Features and Algorithm Results
    # Identify algorithm columns
    algo_cols = ['topsis', 'vikor', 'hybrid', 'rank']
    existing_algos = [c for c in algo_cols if c in df.columns]
    
    # Base columns (everything except algos)
    base_cols = [c for c in df.columns if c not in algo_cols and c != 'id']
    
    df_raw = df[base_cols]
    
    if len(existing_algos) > 0:
        # Results columns
        df_results = df[['Site', 'name'] + existing_algos].copy()
        if 'name' in df_results.columns and 'Site' in df_results.columns:
            df_results = df_results.drop(columns=['name']) # Cleanup duplicate names
    else:
        df_results = pd.DataFrame()

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_raw.to_excel(writer, sheet_name="Raw Dataset", index=False)
        
        if not df_results.empty:
            df_results.to_excel(writer, sheet_name="Algorithm Scores", index=False)
            
        workbook = writer.book
        
        # Style Headers for all sheets
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = (max_length + 2)

        # ---------------------------------------------------------
        # Dedicated Visualizations Dashboard
        # ---------------------------------------------------------
        if not df_results.empty:
            ws_results = workbook["Algorithm Scores"]
            ws_charts = workbook.create_sheet("Executive Visualizations", 0) # Insert as first tab
            workbook.active = ws_charts
            
            # Hide gridlines for a true "Dashboard" feel
            ws_charts.sheet_view.showGridLines = False
            
            # Master Title
            ws_charts['B2'] = "SMART PLANT LOCATION DSS - EXECUTIVE ANALYTICS"
            ws_charts['B2'].font = Font(size=24, bold=True, color="1E3A8A")
            
            # Find column indices
            hybrid_idx, topsis_idx, vikor_idx = None, None, None
            for idx, col in enumerate(df_results.columns):
                c_lower = col.lower()
                if c_lower == 'hybrid': hybrid_idx = idx + 1
                elif c_lower == 'topsis': topsis_idx = idx + 1
                elif c_lower == 'vikor': vikor_idx = idx + 1
                    
            if hybrid_idx:
                # Add Explanatory Text for Hybrid
                ws_charts['B4'] = "1. ULTIMATE HYBRID RANKING DISTRIBUTION"
                ws_charts['B4'].font = Font(size=14, bold=True, color="000000")
                ws_charts['B5'] = "Mathematical Logic: The Hybrid Score computationally synthesizes the optimistic Vector distance (TOPSIS)"
                ws_charts['B6'] = "with the pessimistic structural bound (VIKOR) to locate the absolute best strategic location."
                ws_charts['B5'].font = Font(italic=True, color="4B5563")
                ws_charts['B6'].font = Font(italic=True, color="4B5563")

                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 44 # Corporate professional theme
                chart1.title = "Final Hybrid Ranking"
                chart1.y_axis.title = "Hybrid Score (Higher is Better)"
                chart1.x_axis.title = "Candidate Facilities"
                chart1.legend = None # Remove legend since it's just one metric

                # Data References from Algorithm Scores sheet
                data_ref1 = Reference(ws_results, min_col=hybrid_idx, min_row=1, max_row=len(df_results)+1)
                cats_ref1 = Reference(ws_results, min_col=1, min_row=2, max_row=len(df_results)+1)
                
                chart1.add_data(data_ref1, titles_from_data=True)
                chart1.set_categories(cats_ref1)
                chart1.shape = 4
                chart1.height = 13
                chart1.width = 24
                
                ws_charts.add_chart(chart1, "B8")

            if topsis_idx and vikor_idx:
                # Add Explanatory Text for Divergence
                ws_charts['N4'] = "2. ALGORITHMIC DIVERGENCE (TOPSIS vs VIKOR)"
                ws_charts['N4'].font = Font(size=14, bold=True, color="000000")
                ws_charts['N5'] = "Mathematical Logic: Compares how close a site is to geometric perfection (TOPSIS)"
                ws_charts['N6'] = "against how mathematically safe it is as a compromise penalty choice (VIKOR)."
                ws_charts['N5'].font = Font(italic=True, color="4B5563")
                ws_charts['N6'].font = Font(italic=True, color="4B5563")

                chart2 = BarChart()
                chart2.type = "col"
                chart2.style = 10 # Multi-color theme
                chart2.title = "TOPSIS vs VIKOR Analysis"
                chart2.y_axis.title = "Normalized Index"
                chart2.x_axis.title = "Candidate Facilities"
                
                data_ref2_topsis = Reference(ws_results, min_col=topsis_idx, min_row=1, max_row=len(df_results)+1)
                data_ref2_vikor = Reference(ws_results, min_col=vikor_idx, min_row=1, max_row=len(df_results)+1)
                
                chart2.add_data(data_ref2_topsis, titles_from_data=True)
                chart2.add_data(data_ref2_vikor, titles_from_data=True)
                chart2.set_categories(cats_ref1)
                chart2.height = 13
                chart2.width = 24
                
                ws_charts.add_chart(chart2, "N8") 
                
            # Chart 3: Raw Parameter Breakdown (Cost vs Infrastructure if exists)
            ws_raw = workbook["Raw Dataset"]
            capex_idx, vendor_idx = None, None
            for idx, col in enumerate(df_raw.columns):
                c_lower = col.lower()
                if "capex" in c_lower: capex_idx = idx + 1
                elif "vendor" in c_lower: vendor_idx = idx + 1
                
            if capex_idx and vendor_idx:
                ws_charts['B35'] = "3. RAW CAPABILITY VS COST PROFILE"
                ws_charts['B35'].font = Font(size=14, bold=True, color="000000")
                ws_charts['B36'] = "Dataset Logic: This chart bypasses the algorithms and plots the raw Capital Expenditure against"
                ws_charts['B37'] = "the local Vendor Base infrastructure directly from your uploaded data to visualize trade-offs."
                ws_charts['B36'].font = Font(italic=True, color="4B5563")
                ws_charts['B37'].font = Font(italic=True, color="4B5563")

                chart3 = BarChart()
                chart3.type = "bar" # Horizontal bars for variety
                chart3.style = 27
                chart3.title = "Capex vs Vendor Base Trade-off"
                chart3.x_axis.title = "Raw Output"
                chart3.y_axis.title = "Candidate Facilities"
                
                data_ref_capex = Reference(ws_raw, min_col=capex_idx, min_row=1, max_row=len(df_raw)+1)
                data_ref_vendor = Reference(ws_raw, min_col=vendor_idx, min_row=1, max_row=len(df_raw)+1)
                cats_ref_raw = Reference(ws_raw, min_col=1, min_row=2, max_row=len(df_raw)+1)
                
                chart3.add_data(data_ref_capex, titles_from_data=True)
                chart3.add_data(data_ref_vendor, titles_from_data=True)
                chart3.set_categories(cats_ref_raw)
                chart3.height = 14
                chart3.width = 30
                
                ws_charts.add_chart(chart3, "B39")

    output.seek(0)
    return output

def generate_pdf_summary(summary_data):
    """
    Generate a professional PDF summary report.
    summary_data: {title, description, top_recommendations: []}
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"<b>{summary_data['title']}</b>", styles['Title']))
    elements.append(Spacer(1, 12))
    
    # Description
    elements.append(Paragraph(summary_data['description'], styles['Normal']))
    elements.append(Spacer(1, 24))

    # Table
    table_data = [["Rank", "Location", "Hybrid Score", "Status"]]
    for i, rec in enumerate(summary_data['top_recommendations']):
        table_data.append([str(i+1), rec['name'], f"{rec['score']:.4f}", rec['status']])

    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#EAB308")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer
